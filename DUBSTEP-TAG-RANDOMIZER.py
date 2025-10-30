import random
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import openpyxl
from openpyxl.utils import column_index_from_string
import re

def load_tags_from_file(filename):
    """Загружает теги из файла, обрабатывая как разделенные запятыми, так и построчные теги"""
    try:
        if not os.path.exists(filename):
            return []
        
        with open(filename, 'r', encoding='utf-8') as file:
            content = file.read().strip()
            
        # Разделяем теги по запятым ИЛИ по строкам
        tags = []
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            if ',' in line:
                # Если строка содержит запятые, разделяем по ним
                tags.extend([tag.strip() for tag in line.split(',') if tag.strip()])
            else:
                # Если строка без запятых, используем как отдельный тег
                if line:
                    tags.append(line)
        
        # Удаляем дубликаты
        tags = list(dict.fromkeys(tags))
        return tags
    
    except Exception as e:
        print(f"❌ Ошибка при чтении файла {filename}: {e}")
        return []

def save_tags_to_file(tags, filename):
    """Сохраняет теги в файл через запятую"""
    try:
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(', '.join(tags))
        return True
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        return False

def parse_excel_cell(cell_address):
    """Парсит адрес ячейки Excel (например, 'D4') на строку и столбец"""
    try:
        # Регулярное выражение для поиска букв (столбец) и цифр (строка)
        match = re.match(r'([A-Za-z]+)(\d+)$', cell_address.upper())
        if match:
            col_letter = match.group(1)
            row_number = int(match.group(2))
            return row_number, col_letter
        return None, None
    except:
        return None, None

def write_to_excel(tags_list, filename, start_cell, direction, sheet_name=None):
    """Записывает список тегов в Excel файл"""
    try:
        # Пытаемся загрузить существующий файл или создать новый
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
        else:
            workbook = openpyxl.Workbook()
        
        # Выбираем лист
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
        else:
            sheet = workbook.active
        
        # Парсим начальную ячейку
        start_row, start_col_letter = parse_excel_cell(start_cell)
        if not start_row or not start_col_letter:
            return False, "error_excel_cell"
        
        start_col = column_index_from_string(start_col_letter)
        
        # Записываем теги
        current_row = start_row
        current_col = start_col
        
        for tags in tags_list:
            if direction == "down":
                sheet.cell(row=current_row, column=current_col).value = tags
                current_row += 1
            else:  # right
                sheet.cell(row=current_row, column=current_col).value = tags
                current_col += 1
        
        # Сохраняем файл
        workbook.save(filename)
        return True, "success_excel_write"
        
    except Exception as e:
        return False, f"error_excel_write_generic: {str(e)}"

class TagGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🏷️ Генератор тегов")
        self.root.geometry("680x700")
        self.root.resizable(True, True)
        
        # Текущий язык (по умолчанию русский)
        self.current_language = "RU"
        
        # Тексты для перевода
        self.translations = {
            "RU": {
                "title": "🏷️ Генератор тегов",
                "available_tags": "📊 Доступные теги",
                "main_tags": "Главные теги: {} штук",
                "important_tags": "Важные теги: {} штук",
                "additional_tags": "Остальные теги: {} штук",
                "total_tags_frame": "1. Общее количество тегов",
                "total_tags_label": "Сколько всего тегов должно быть сформировано?",
                "important_tags_frame": "2. Важные теги",
                "important_tags_label": "Сколько важных тегов добавить?",
                "other_tags_frame": "3. Остальные теги",
                "other_tags_label": "Сколько остальных тегов будет добавлено:",
                "excel_frame": "📊 Запись в Excel (опционально)",
                "excel_check": "Записать результаты в Excel файл",
                "excel_file": "Файл Excel:",
                "excel_sheet": "Лист:",
                "excel_cell": "Начальная ячейка:",
                "excel_cell_example": "(например: D4)",
                "excel_direction": "Направление:",
                "excel_down": "Вниз (столбец)",
                "excel_right": "Вправо (строка)",
                "excel_generations": "Количество генераций:",
                "status_ready": "Готов к работе",
                "status_enter_total": "Введите общее количество тегов",
                "status_too_many_main": "❌ Слишком много главных тегов! Уменьшите общее количество или главные теги",
                "status_too_many_important": "❌ Слишком много важных тегов!",
                "status_not_enough_important": "❌ Запрошено больше важных тегов чем доступно!",
                "status_ready_to_generate": "✅ Всё готово к генерации!",
                "status_enter_valid_numbers": "⚠️ Введите корректные числа",
                "generate_button": "🚀 Сгенерировать теги",
                "result_frame": "📋 Результат",
                "copy_button": "📋 Копировать результат",
                "language_button": "ENG",
                "error_title": "Ошибка",
                "success_title": "Успех",
                "warning_title": "Внимание",
                "error_main_tags_exceed": "Количество главных тегов превышает общее количество!",
                "error_too_many_important": "Слишком много важных тегов!",
                "error_not_enough_important": "Запрошено больше важных тегов чем доступно!",
                "error_generations_positive": "Количество генераций должно быть положительным числом!",
                "error_enter_valid_numbers": "Пожалуйста, введите корректные числа!",
                "error_excel_cell": "Неверный формат ячейки. Используйте формат как 'D4'",
                "error_excel_write": "Ошибка при записи в Excel: {}",
                "error_save_file": "Не удалось сохранить файл!",
                "error_unknown": "Произошла ошибка: {}",
                "success_generated_single": "✅ Сгенерировано {} тегов!\n📁 Сохранено в: result_tags.txt\n\n📊 Статистика:\n• Главные теги: {}\n• Важные теги: {}\n• Остальные теги: {}",
                "success_generated_multiple": "✅ Сгенерировано {} наборов тегов!\n📊 Записано в Excel файл\n\n{}",
                "success_copied": "Результат скопирован в буфер обмена!",
                "warning_no_result": "Нет результата для копирования!",
                "success_excel_message": "Успешно записано {} наборов тегов в файл {}",
                "error_excel_write_generic": "Ошибка при записи в Excel: {}"
            },
            "ENG": {
                "title": "🏷️ Tag Generator",
                "available_tags": "📊 Available Tags",
                "main_tags": "Main tags: {} items",
                "important_tags": "Important tags: {} items",
                "additional_tags": "Other tags: {} items",
                "total_tags_frame": "1. Total Number of Tags",
                "total_tags_label": "How many total tags should be generated?",
                "important_tags_frame": "2. Important Tags",
                "important_tags_label": "How many important tags to add?",
                "other_tags_frame": "3. Other Tags",
                "other_tags_label": "How many other tags will be added:",
                "excel_frame": "📊 Excel Export (optional)",
                "excel_check": "Save results to Excel file",
                "excel_file": "Excel file:",
                "excel_sheet": "Sheet:",
                "excel_cell": "Starting cell:",
                "excel_cell_example": "(e.g., D4)",
                "excel_direction": "Direction:",
                "excel_down": "Down (column)",
                "excel_right": "Right (row)",
                "excel_generations": "Number of generations:",
                "status_ready": "Ready to work",
                "status_enter_total": "Enter total number of tags",
                "status_too_many_main": "❌ Too many main tags! Reduce total number or main tags",
                "status_too_many_important": "❌ Too many important tags!",
                "status_not_enough_important": "❌ Requested more important tags than available!",
                "status_ready_to_generate": "✅ Everything ready to generate!",
                "status_enter_valid_numbers": "⚠️ Enter valid numbers",
                "generate_button": "🚀 Generate Tags",
                "result_frame": "📋 Result",
                "copy_button": "📋 Copy Result",
                "language_button": "RU",
                "error_title": "Error",
                "success_title": "Success",
                "warning_title": "Warning",
                "error_main_tags_exceed": "Number of main tags exceeds total number!",
                "error_too_many_important": "Too many important tags!",
                "error_not_enough_important": "Requested more important tags than available!",
                "error_generations_positive": "Number of generations must be a positive number!",
                "error_enter_valid_numbers": "Please enter valid numbers!",
                "error_excel_cell": "Invalid cell format. Use format like 'D4'",
                "error_excel_write": "Error writing to Excel: {}",
                "error_save_file": "Failed to save file!",
                "error_unknown": "An error occurred: {}",
                "success_generated_single": "✅ Generated {} tags!\n📁 Saved to: result_tags.txt\n\n📊 Statistics:\n• Main tags: {}\n• Important tags: {}\n• Other tags: {}",
                "success_generated_multiple": "✅ Generated {} sets of tags!\n📊 Written to Excel file\n\n{}",
                "success_copied": "Result copied to clipboard!",
                "warning_no_result": "No result to copy!",
                "success_excel_message": "Successfully written {} sets of tags to file {}",
                "error_excel_write_generic": "Error writing to Excel: {}"
            }
        }
        
        # Центрируем окно
        self.center_window()
        
        # Загружаем теги при запуске
        self.permanent_tags = load_tags_from_file('main_tags.txt')
        self.important_tags = load_tags_from_file('additional_tags.txt')
        self.additional_tags = load_tags_from_file('other_tags.txt')
        
        self.setup_ui()
        self.update_counts()
    
    def center_window(self):
        """Центрирует окно на экране"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def get_text(self, key):
        """Возвращает текст на текущем языке"""
        return self.translations[self.current_language].get(key, key)
    
    def toggle_language(self):
        """Переключает язык интерфейса"""
        self.current_language = "ENG" if self.current_language == "RU" else "RU"
        self.update_ui_language()
    
    def update_ui_language(self):
        """Обновляет все тексты интерфейса на текущем языке"""
        # Обновляем заголовок окна
        self.root.title(self.get_text("title"))
        
        # Обновляем кнопку языка
        self.language_btn.config(text=self.get_text("language_button"))
        
        # Обновляем заголовки фреймов
        self.title_label.config(text=self.get_text("title"))
        self.info_frame.config(text=self.get_text("available_tags"))
        self.total_frame.config(text=self.get_text("total_tags_frame"))
        self.important_frame.config(text=self.get_text("important_tags_frame"))
        self.other_frame.config(text=self.get_text("other_tags_frame"))
        self.excel_file_label.config(text=self.get_text("excel_file"))
        self.sheet_label.config(text=self.get_text("excel_sheet"))
        self.cell_label.config(text=self.get_text("excel_cell"))
        self.cell_example_label.config(text=self.get_text("excel_cell_example"))
        self.direction_label.config(text=self.get_text("excel_direction"))
        self.down_radio.config(text=self.get_text("excel_down"))
        self.right_radio.config(text=self.get_text("excel_right"))
        self.generations_label.config(text=self.get_text("excel_generations"))
        self.result_frame.config(text=self.get_text("result_frame"))
        
        # Обновляем тексты меток
        self.total_label.config(text=self.get_text("total_tags_label"))
        self.important_label.config(text=self.get_text("important_tags_label"))
        self.other_label.config(text=self.get_text("other_tags_label"))
        self.excel_check.config(text=self.get_text("excel_check"))
        
        # Обновляем кнопки
        self.generate_btn.config(text=self.get_text("generate_button"))
        self.copy_btn.config(text=self.get_text("copy_button"))
        
        # Обновляем статус
        self.update_counts()
        self.update_status()
    
    def setup_ui(self):
        # Устанавливаем фон для корневого окна
        self.root.configure(bg='#f0f0f0')
        
        # Создаем главный контейнер с прокруткой
        main_container = tk.Frame(self.root, bg='#f0f0f0')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Создаем canvas и скроллбар
        self.canvas = tk.Canvas(main_container, bg='#f0f0f0', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='#f0f0f0')
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        # Создаем окно в canvas
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # Упаковываем canvas и скроллбар
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Привязываем колесо мыши к прокрутке
        self._bind_to_mousewheel(self.scrollable_frame)
        
        # Обработка изменения размера окна
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        
        main_frame = tk.Frame(self.scrollable_frame, bg='#f0f0f0', padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Заголовок с кнопкой языка
        title_frame = tk.Frame(main_frame, bg='#f0f0f0')
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.title_label = tk.Label(title_frame, text=self.get_text("title"), 
                              font=('Arial', 12, 'bold'), bg='#f0f0f0')
        self.title_label.pack(side=tk.LEFT)
        
        # Кнопка переключения языка в правом верхнем углу
        self.language_btn = tk.Button(title_frame, text=self.get_text("language_button"),
                                    font=('Arial', 9), width=4, height=1,
                                    command=self.toggle_language, relief='raised', bd=1)
        self.language_btn.pack(side=tk.RIGHT)
        
        # Информация о доступных тегах
        self.info_frame = tk.LabelFrame(main_frame, text=self.get_text("available_tags"), bg='#f0f0f0', 
                                 font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.info_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.permanent_count_label = tk.Label(self.info_frame, text="", 
                                            font=('Arial', 9), bg='#f0f0f0', anchor='w')
        self.permanent_count_label.pack(fill=tk.X)
        
        self.important_count_label = tk.Label(self.info_frame, text="", 
                                            font=('Arial', 9), bg='#f0f0f0', anchor='w')
        self.important_count_label.pack(fill=tk.X)
        
        self.additional_count_label = tk.Label(self.info_frame, text="", 
                                             font=('Arial', 9), bg='#f0f0f0', anchor='w')
        self.additional_count_label.pack(fill=tk.X)
        
        # Поле для общего количества тегов
        self.total_frame = tk.LabelFrame(main_frame, text=self.get_text("total_tags_frame"), bg='#f0f0f0',
                                  font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.total_frame.pack(fill=tk.X, pady=(0, 8))
        
        total_inner_frame = tk.Frame(self.total_frame, bg='#f0f0f0')
        total_inner_frame.pack(fill=tk.X)
        
        # Текст total_tags
        self.total_label = tk.Label(total_inner_frame, text=self.get_text("total_tags_label"), 
                font=('Arial', 9), bg='#f0f0f0', width=42, anchor='w', justify='left')
        self.total_label.pack(side=tk.LEFT, fill='x', expand=True)
        
        self.total_var = tk.StringVar()
        self.total_entry = tk.Entry(total_inner_frame, textvariable=self.total_var, 
                                  font=('Arial', 10), width=8, relief='solid', bd=1)
        self.total_entry.pack(side=tk.RIGHT, padx=(10, 0))
        
        self.total_var.trace('w', self.on_total_changed)
        
        # Поле для важных тегов
        self.important_frame = tk.LabelFrame(main_frame, text=self.get_text("important_tags_frame"), bg='#f0f0f0',
                                      font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.important_frame.pack(fill=tk.X, pady=(0, 8))
        
        important_inner_frame = tk.Frame(self.important_frame, bg='#f0f0f0')
        important_inner_frame.pack(fill=tk.X)
        
        # Текст important_tags
        self.important_label = tk.Label(important_inner_frame, text=self.get_text("important_tags_label"), 
                font=('Arial', 9), bg='#f0f0f0', width=42, anchor='w', justify='left')
        self.important_label.pack(side=tk.LEFT, fill='x', expand=True)
        
        self.important_var = tk.StringVar()
        self.important_entry = tk.Entry(important_inner_frame, textvariable=self.important_var, 
                                      font=('Arial', 10), width=8, relief='solid', bd=1)
        self.important_entry.pack(side=tk.RIGHT, padx=(10, 0))
        
        self.important_var.trace('w', self.on_important_changed)
        
        # Поле для остальных тегов (только чтение)
        self.other_frame = tk.LabelFrame(main_frame, text=self.get_text("other_tags_frame"), bg='#f0f0f0',
                                  font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.other_frame.pack(fill=tk.X, pady=(0, 10))
        
        other_inner_frame = tk.Frame(self.other_frame, bg='#f0f0f0')
        other_inner_frame.pack(fill=tk.X)
        
        # Текст other_tags
        self.other_label = tk.Label(other_inner_frame, text=self.get_text("other_tags_label"), 
                font=('Arial', 9), bg='#f0f0f0', width=42, anchor='w', justify='left')
        self.other_label.pack(side=tk.LEFT, fill='x', expand=True)
        
        self.other_var = tk.StringVar(value="0")
        self.other_entry = tk.Entry(other_inner_frame, textvariable=self.other_var, 
                                  state='readonly', font=('Arial', 10), foreground='blue', 
                                  width=8, relief='solid', bd=1)
        self.other_entry.pack(side=tk.RIGHT, padx=(10, 0))
        
        # Настройки Excel
        self.excel_frame = tk.LabelFrame(main_frame, text=self.get_text("excel_frame"), bg='#f0f0f0',
                                  font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Чекбокс для включения записи в Excel
        self.excel_var = tk.BooleanVar()
        self.excel_check = tk.Checkbutton(self.excel_frame, text=self.get_text("excel_check"), 
                                        variable=self.excel_var, command=self.toggle_excel_fields,
                                        font=('Arial', 9), bg='#f0f0f0', anchor='w')
        self.excel_check.pack(fill=tk.X, pady=(0, 8))
        
        # Поля для Excel (изначально скрыты)
        self.excel_fields_frame = tk.Frame(self.excel_frame, bg='#f0f0f0')
        
        # Файл Excel
        file_frame = tk.Frame(self.excel_fields_frame, bg='#f0f0f0')
        file_frame.pack(fill=tk.X, pady=(0, 5))
        self.excel_file_label = tk.Label(file_frame, text=self.get_text("excel_file"), font=('Arial', 9), bg='#f0f0f0', width=19, anchor='w', justify='left')
        self.excel_file_label.pack(side=tk.LEFT)
        self.excel_file_var = tk.StringVar(value="tags.xlsx")
        self.excel_file_entry = tk.Entry(file_frame, textvariable=self.excel_file_var, 
                                    font=('Arial', 9), width=15, relief='solid', bd=1)
        self.excel_file_entry.pack(side=tk.LEFT, padx=(10, 0))

        # Лист
        sheet_frame = tk.Frame(self.excel_fields_frame, bg='#f0f0f0')
        sheet_frame.pack(fill=tk.X, pady=(0, 5))
        self.sheet_label = tk.Label(sheet_frame, text=self.get_text("excel_sheet"), font=('Arial', 9), bg='#f0f0f0', width=19, anchor='w', justify='left')
        self.sheet_label.pack(side=tk.LEFT)
        self.sheet_var = tk.StringVar(value="Tags")
        self.sheet_entry = tk.Entry(sheet_frame, textvariable=self.sheet_var, 
                                font=('Arial', 9), width=10, relief='solid', bd=1)
        self.sheet_entry.pack(side=tk.LEFT, padx=(10, 0))

        # Начальная ячейка
        cell_frame = tk.Frame(self.excel_fields_frame, bg='#f0f0f0')
        cell_frame.pack(fill=tk.X, pady=(0, 5))
        self.cell_label = tk.Label(cell_frame, text=self.get_text("excel_cell"), font=('Arial', 9), bg='#f0f0f0', width=19, anchor='w', justify='left')
        self.cell_label.pack(side=tk.LEFT)
        self.cell_var = tk.StringVar(value="A1")
        self.cell_entry = tk.Entry(cell_frame, textvariable=self.cell_var, 
                                font=('Arial', 9), width=8, relief='solid', bd=1)
        self.cell_entry.pack(side=tk.LEFT, padx=(10, 0))
        self.cell_example_label = tk.Label(cell_frame, text=self.get_text("excel_cell_example"), font=('Arial', 8), bg='#f0f0f0', 
                fg='#666666')
        self.cell_example_label.pack(side=tk.LEFT, padx=(5, 0))

        # Направление записи
        direction_frame = tk.Frame(self.excel_fields_frame, bg='#f0f0f0')
        direction_frame.pack(fill=tk.X, pady=(0, 5))
        self.direction_label = tk.Label(direction_frame, text=self.get_text("excel_direction"), font=('Arial', 9), bg='#f0f0f0', width=19, anchor='w', justify='left')
        self.direction_label.pack(side=tk.LEFT)
        self.direction_var = tk.StringVar(value="down")
        self.down_radio = tk.Radiobutton(direction_frame, text=self.get_text("excel_down"), font=('Arial', 9),
                    variable=self.direction_var, value="down", bg='#f0f0f0')
        self.down_radio.pack(side=tk.LEFT, padx=(10, 5))
        self.right_radio = tk.Radiobutton(direction_frame, text=self.get_text("excel_right"), font=('Arial', 9),
                    variable=self.direction_var, value="right", bg='#f0f0f0')
        self.right_radio.pack(side=tk.LEFT)

        # Количество генераций
        count_frame = tk.Frame(self.excel_fields_frame, bg='#f0f0f0')
        count_frame.pack(fill=tk.X, pady=(0, 5))
        self.generations_label = tk.Label(count_frame, text=self.get_text("excel_generations"), font=('Arial', 9), bg='#f0f0f0', width=19, anchor='w', justify='left')
        self.generations_label.pack(side=tk.LEFT)
        self.generations_var = tk.StringVar(value="1")
        self.generations_entry = tk.Entry(count_frame, textvariable=self.generations_var, 
                                        font=('Arial', 9), width=8, relief='solid', bd=1)
        self.generations_entry.pack(side=tk.LEFT, padx=(10, 0))
        
        self.excel_fields_frame.pack(fill=tk.X)
        self.excel_fields_frame.pack_forget()  # Скрываем поля по умолчанию
        
        # Статус
        self.status_var = tk.StringVar(value=self.get_text("status_ready"))
        self.status_label = tk.Label(main_frame, textvariable=self.status_var, 
                                   font=('Arial', 9), bg='#f0f0f0', fg='green')
        self.status_label.pack(pady=(0, 8))
        
        # Кнопка генерации
        self.generate_btn = tk.Button(main_frame, text=self.get_text("generate_button"), 
                                    font=('Arial', 10, 'bold'), bg='#28a745', fg='white',
                                    relief='raised', bd=2, command=self.generate_tags)
        self.generate_btn.pack(pady=(0, 10))
        
        # Область для вывода результата
        self.result_frame = tk.LabelFrame(main_frame, text=self.get_text("result_frame"), bg='#f0f0f0',
                                   font=('Arial', 10, 'bold'), padx=10, pady=10)
        self.result_frame.pack(fill=tk.BOTH, expand=True)
        
        self.result_text = scrolledtext.ScrolledText(self.result_frame, height=8, font=('Arial', 9),
                                                   relief='solid', bd=1)
        self.result_text.pack(fill=tk.BOTH, expand=True)
        
        # Кнопка копирования
        self.copy_btn = tk.Button(main_frame, text=self.get_text("copy_button"), font=('Arial', 9),
                 command=self.copy_result, relief='raised', bd=1)
        self.copy_btn.pack(pady=(8, 0))
    
    def _bind_to_mousewheel(self, widget):
        """Привязывает колесо мыши к виджету для прокрутки"""
        widget.bind("<MouseWheel>", self._on_mousewheel)
        # Привязываем ко всем дочерним элементам
        for child in widget.winfo_children():
            child.bind("<MouseWheel>", self._on_mousewheel)
            if hasattr(child, 'winfo_children') and child.winfo_children():
                self._bind_to_mousewheel(child)
    
    def _on_mousewheel(self, event):
        """Обработчик прокрутки колесом мыши"""
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def on_canvas_configure(self, event):
        """Обработчик изменения размера canvas"""
        # Устанавливаем ширину прокручиваемой области равной ширине canvas
        self.canvas.itemconfig(self.canvas_frame, width=event.width)
    
    def toggle_excel_fields(self):
        """Показывает/скрывает поля для Excel"""
        if self.excel_var.get():
            self.excel_fields_frame.pack(fill=tk.X)
        else:
            self.excel_fields_frame.pack_forget()
    
    def update_counts(self):
        """Обновляет информацию о доступных тегах"""
        self.permanent_count_label.config(
            text=self.get_text("main_tags").format(len(self.permanent_tags))
        )
        self.important_count_label.config(
            text=self.get_text("important_tags").format(len(self.important_tags))
        )
        self.additional_count_label.config(
            text=self.get_text("additional_tags").format(len(self.additional_tags))
        )
    
    def on_total_changed(self, *args):
        """Обработчик изменения общего количества тегов"""
        self.update_other_count()
        self.update_status()
    
    def on_important_changed(self, *args):
        """Обработчик изменения количества важных тегов"""
        self.update_other_count()
        self.update_status()
    
    def update_other_count(self):
        """Обновляет количество остальных тегов"""
        try:
            total = int(self.total_var.get()) if self.total_var.get() else 0
            important = int(self.important_var.get()) if self.important_var.get() else 0
            
            if total > 0:
                permanent_used = min(len(self.permanent_tags), total)
                available_for_other = total - permanent_used - important
                other_count = max(0, min(available_for_other, len(self.additional_tags)))
                self.other_var.set(str(other_count))
            else:
                self.other_var.set("0")
        except ValueError:
            self.other_var.set("0")
    
    def update_status(self):
        """Обновляет статусную строку"""
        try:
            total = int(self.total_var.get()) if self.total_var.get() else 0
            important = int(self.important_var.get()) if self.important_var.get() else 0
            
            if total == 0:
                self.status_var.set(self.get_text("status_enter_total"))
                self.status_label.config(fg='black')
                return
            
            # Проверка главных тегов
            if len(self.permanent_tags) > total:
                self.status_var.set(self.get_text("status_too_many_main"))
                self.status_label.config(fg='red')
                return
            
            permanent_used = min(len(self.permanent_tags), total)
            
            # Проверка важных тегов
            if important > (total - permanent_used):
                self.status_var.set(self.get_text("status_too_many_important"))
                self.status_label.config(fg='red')
                return
            
            if important > len(self.important_tags):
                self.status_var.set(self.get_text("status_not_enough_important"))
                self.status_label.config(fg='red')
                return
            
            self.status_var.set(self.get_text("status_ready_to_generate"))
            self.status_label.config(fg='green')
            
        except ValueError:
            self.status_var.set(self.get_text("status_enter_valid_numbers"))
            self.status_label.config(fg='orange')
    
    def generate_single_tags(self):
        """Генерирует один набор тегов"""
        total_tags_needed = int(self.total_var.get())
        important_count = int(self.important_var.get()) if self.important_var.get() else 0
        
        # Формирование результата
        result_tags = self.permanent_tags.copy()
        remaining_slots = total_tags_needed - len(result_tags)
        
        # Добавление важных теги
        if remaining_slots > 0 and important_count > 0:
            selected_important = random.sample(self.important_tags, important_count)
            result_tags.extend(selected_important)
            remaining_slots -= important_count
        
        # Добавление остальных тегов
        if remaining_slots > 0 and self.additional_tags:
            additional_count = min(remaining_slots, len(self.additional_tags))
            if additional_count > 0:
                selected_additional = random.sample(self.additional_tags, additional_count)
                result_tags.extend(selected_additional)
        
        # Обрезаем если нужно
        if len(result_tags) > total_tags_needed:
            result_tags = result_tags[:total_tags_needed]
        
        return ', '.join(result_tags)
    
    def generate_tags(self):
        """Генерация тегов"""
        try:
            total_tags_needed = int(self.total_var.get())
            important_count = int(self.important_var.get()) if self.important_var.get() else 0
            
            # Проверки
            if len(self.permanent_tags) > total_tags_needed:
                messagebox.showerror(self.get_text("error_title"), self.get_text("error_main_tags_exceed"))
                return
            
            if important_count > (total_tags_needed - len(self.permanent_tags)):
                messagebox.showerror(self.get_text("error_title"), self.get_text("error_too_many_important"))
                return
            
            if important_count > len(self.important_tags):
                messagebox.showerror(self.get_text("error_title"), self.get_text("error_not_enough_important"))
                return
            
            # Генерация тегов
            if self.excel_var.get():
                # Множественная генерация для Excel
                generations = int(self.generations_var.get()) if self.generations_var.get() else 1
                if generations <= 0:
                    messagebox.showerror(self.get_text("error_title"), self.get_text("error_generations_positive"))
                    return
                
                tags_list = []
                for i in range(generations):
                    tags_str = self.generate_single_tags()
                    tags_list.append(tags_str)
                
                # Запись в Excel
                success, message = write_to_excel(
                    tags_list,
                    self.excel_file_var.get(),
                    self.cell_var.get(),
                    self.direction_var.get(),
                    self.sheet_var.get() if self.sheet_var.get() else None
                )

                if success:
                    # Показываем первый результат
                    self.result_text.delete(1.0, tk.END)
                    if self.current_language == "RU":
                        self.result_text.insert(1.0, f"Сгенерировано {generations} наборов тегов:\n\n")
                    else:
                        self.result_text.insert(1.0, f"Generated {generations} sets of tags:\n\n")
                    
                    for i, tags in enumerate(tags_list[:5], 1):
                        self.result_text.insert(tk.END, f"{i}. {tags}\n\n")
                    
                    if generations > 5:
                        if self.current_language == "RU":
                            self.result_text.insert(tk.END, f"... и ещё {generations - 5} наборов\n")
                        else:
                            self.result_text.insert(tk.END, f"... and {generations - 5} more sets\n")
                    
                    # Формируем сообщение успеха
                    if message == "success_excel_write":
                        success_message = self.get_text("success_generated_multiple").format(
                            generations, 
                            self.get_text("success_excel_message").format(len(tags_list), self.excel_file_var.get())
                        )
                    else:
                        success_message = self.get_text("success_generated_multiple").format(generations, message)
                    
                    messagebox.showinfo(self.get_text("success_title"), success_message)
                else:
                    # Обработка ошибок
                    if message.startswith("error_excel_write_generic:"):
                        error_detail = message.split(":", 1)[1]
                        error_msg = self.get_text("error_excel_write").format(error_detail)
                    else:
                        error_msg = self.get_text(message) if message in self.translations[self.current_language] else message
                    
                    messagebox.showerror(self.get_text("error_title"), error_msg)
            else:
                # ОДИНОЧНАЯ генерация (когда Excel не выбран)
                tags_str = self.generate_single_tags()
                
                # Сохраняем в файл
                if save_tags_to_file(tags_str.split(', '), 'result_tags.txt'):
                    # Показываем результат
                    self.result_text.delete(1.0, tk.END)
                    self.result_text.insert(1.0, tags_str)
                    
                    # Подсчитываем статистику
                    tags_count = len(tags_str.split(', '))
                    permanent_used = min(len(self.permanent_tags), total_tags_needed)
                    other_used = tags_count - permanent_used - important_count
                    
                    messagebox.showinfo(self.get_text("success_title"), 
                        self.get_text("success_generated_single").format(
                            tags_count, 
                            permanent_used,
                            important_count,
                            other_used
                        )
                    )
                else:
                    messagebox.showerror(self.get_text("error_title"), self.get_text("error_save_file"))
                
        except ValueError:
            messagebox.showerror(self.get_text("error_title"), self.get_text("error_enter_valid_numbers"))
        except Exception as e:
            error_msg = self.get_text("error_unknown").format(str(e))
            messagebox.showerror(self.get_text("error_title"), error_msg)
    
    def copy_result(self):
        """Копирование результата в буфер обмена"""
        result = self.result_text.get(1.0, tk.END).strip()
        if result:
            self.root.clipboard_clear()
            self.root.clipboard_append(result)
            messagebox.showinfo(self.get_text("success_title"), self.get_text("success_copied"))
        else:
            messagebox.showwarning(self.get_text("warning_title"), self.get_text("warning_no_result"))

def main():
    # Запуск GUI приложения
    root = tk.Tk()
    app = TagGeneratorApp(root)
    root.mainloop()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ Произошла ошибка: {e}")
        input("Нажмите Enter для выхода...")